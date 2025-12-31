"""XLSX/XLS parser for ParseForge."""

import logging
from pathlib import Path

from openpyxl import load_workbook

from src.schema.document import BlockType, Document, Page, TableBlock
from src.utils.exceptions import ParserError

logger = logging.getLogger(__name__)


def parse_xlsx(file_path: str) -> Document:
    """
    Parse XLSX file.

    Args:
        file_path: Path to XLSX file

    Returns:
        Document object
    """
    try:
        wb = load_workbook(file_path)
        pages = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]

            # Convert sheet to cells
            cells = []
            for row in sheet.iter_rows(values_only=True):
                cells.append([str(cell) if cell is not None else "" for cell in row])

            if cells:
                # Create table block
                table = TableBlock(
                    block_type=BlockType.TABLE,
                    page_index=sheet_idx,
                    cells=cells,
                    num_rows=len(cells),
                    num_cols=len(cells[0]) if cells else 0,
                )

                # Create page
                page = Page(
                    page_index=sheet_idx,
                    width=612,
                    height=792,
                    blocks=[table],
                )
                pages.append(page)

        document = Document(
            file_path=file_path,
            file_name=Path(file_path).name,
            pages=pages,
            total_pages=len(pages),
        )

        return document

    except Exception as e:
        raise ParserError(f"Failed to parse XLSX: {e}") from e

